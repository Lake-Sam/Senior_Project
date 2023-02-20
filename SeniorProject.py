import stocker 

import tensorflow as tf
import pandas as pd
import statistics
import sys
import time
import os                                               #Importing all the librarys and such I think I may need.
import yfinance as yf
import smtplib, ssl
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
import sharepy                                         #Importing the library to integrate with sharepoint lists
from sharepy import connect
from sharepy import SharePointSession
import requests
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


import concurrent.futures
from numba import jit, cuda
import numpy as np

import tkinter as tk                                    #GUI


from collections import OrderedDict


import datetime as dt
from datetime import date, timedelta
from pytz import timezone


from requests.exceptions import ConnectionError
import logging
tf.get_logger().setLevel(logging.ERROR)

fivehuna = ['AAPL', 'TSLA', 'ABT', 'ABBV', 'ACN', 'ADBE', 'ADT', 'AAP', 'AES', 'AFL', 'AMG', 'A', 'GAS', 'APD', 'AKAM', 'AA', 'ALLE', 'ALL', 'ALTR', 'MO', 'AMZN', 'AEE', 'AAL', 'AEP', 'AXP', 'AIG', 'AMT', 'AMP', 'ABC', 'AME', 'AMGN', 'APH', 'ADI', 'AON', 'APA', 'AIV', 'AMAT', 'ADM', 'AIZ', 'T', 'ADSK', 'ADP', 'AN', 'AZO', 'AVGO', 'AVB', 'AVY', 'BAC', 'BK', 'BAX', 'BDX', 'BBBY', 'BRK-B', 'BBY', 'BLX', 'HRB', 'BA', 'BWA', 'BXP', 'BMY', 'BF-B', 'CHRW', 'CPB', 'COF', 'CAH', 'HSIC', 'KMX', 'CCL', 'CAT', 'CNP', 'CERN', 'CF', 'SCHW', 'CVX', 'CMG', 'CB', 'CI', 'CINF', 'CTAS', 'CSCO', 'C', 'CTXS', 'CLX', 'CME', 'CMS', 'KO', 'CTSH', 'CL', 'CMCSA', 'CMA', 'CSC', 'CAG', 'COP', 'CNX', 'ED', 'STZ', 'GLW', 'COST', 'CCI', 'CSX', 'CMI', 'CVS', 'DHI', 'DHR', 'DRI', 'DVA', 'DE', 'DAL', 'XRAY', 'DVN',  'DFS', 'DG', 'DLTR', 'D', 'DOV', 'DOW', 'DTE', 'DD', 'DUK', 'DNB', 'EMN', 'ETN', 'EBAY', 'ECL', 'EIX', 'EW', 'EA', 'EMC', 'EMR', 'ENDP', 'ETR', 'EOG', 'EQT', 'EFX', 'EQIX', 'EQR', 'ESS', 'EL', 'ES', 'EXC', 'EXPE', 'EXPD', 'XOM', 'FFIV', 'FB', 'FAST', 'FDX', 'FIS', 'FITB', 'FSLR', 'FE', 'FLS', 'FLR', 'FMC', 'FTI', 'F', 'FOSL', 'BEN', 'FCX', 'GME', 'GPS', 'GRMN', 'GD', 'GE', 'GIS', 'GM', 'GPC', 'GNW', 'GILD', 'GS', 'GT', 'GOOGL', 'GOOG', 'GWW', 'HAL', 'HBI', 'HOG', 'HAR', 'HIG', 'HAS', 'HCA', 'HP', 'HES', 'HPQ', 'HD', 'HON', 'HRL', 'HST', 'HUM', 'HBAN', 'ITW', 'IR', 'INTC', 'ICE', 'IBM', 'IP', 'IPG', 'IFF', 'INTU', 'ISRG', 'IVZ', 'IRM', 'JBHT', 'JNJ', 'JCI', 'JPM', 'JNPR', 'K', 'KEY', 'KMB', 'KIM', 'KMI', 'KLAC', 'KSS',  'KR', 'LH', 'LRCX', 'LEG', 'LEN', 'LLY', 'LNC', 'LMT', 'L', 'LOW', 'LYB', 'MTB', 'MAC', 'M', 'MRO', 'MPC', 'MAR', 'MMC', 'MLM', 'MAS', 'MA', 'MAT', 'MKC', 'MCD', 'MCK', 'MDT', 'MRK', 'MET', 'MCHP', 'MU', 'MSFT', 'MHK', 'TAP', 'MDLZ', 'MNST', 'MCO', 'MS', 'MOS', 'MSI', 'MUR', 'NDAQ', 'NOV', 'NAVI', 'NTAP', 'NFLX', 'NWL', 'NEM', 'NWSA', 'NEE', 'NLSN', 'NKE', 'NI', 'JWN', 'NSC', 'NTRS', 'NOC', 'NRG', 'NUE', 'NVDA', 'ORLY', 'OXY', 'OMC', 'OKE', 'ORCL', 'OI', 'PCAR', 'PLL', 'PH', 'PDCO', 'PAYX', 'PNR', 'POM', 'PEP', 'PKI', 'PRGO', 'PFE', 'PCG', 'PM', 'PSX', 'PNW', 'PXD', 'PBI', 'PCL', 'PNC', 'RL', 'PPG', 'PPL', 'PFG', 'PG', 'PGR', 'PLD', 'PRU', 'PEG', 'PSA', 'PHM', 'PVH', 'QRVO', 'PWR', 'QCOM', 'DGX', 'RRC', 'O', 'REGN', 'RF', 'RSG', 'RHI', 'ROK', 'ROP', 'ROST', 'R', 'CRM', 'SLB', 'STX', 'SEE', 'SRE', 'SHW', 'SPG', 'SWKS', 'SLG', 'SJM', 'SNA', 'SO', 'LUV', 'SWN', 'SE', 'SWK', 'SBUX', 'STT', 'SRCL', 'SYK', 'SYY', 'TROW', 'TGT', 'TEL', 'TGNA', 'THC', 'TDC', 'TXN', 'TXT', 'HSY', 'TRV', 'TMO', 'TSCO', 'RIG', 'TRIP', 'FOXA', 'TSN', 'UA', 'UNP', 'UNH', 'UPS', 'URI', 'UHS', 'UNM', 'URBN', 'VFC', 'VLO', 'VTR', 'VRSN', 'VZ', 'VRTX', 'V', 'VNO', 'VMC', 'WMT', 'WBA', 'DIS', 'WM', 'WAT', 'ANTM', 'WFC', 'WDC', 'WU', 'WY', 'WHR', 'WMB', 'WEC', 'WYNN', 'XEL', 'XRX', 'XL', 'XYL', 'YUM', 'ZBH', 'ZION', 'ZTS']
pe = [None] * len(fivehuna)             #profit equity
beta = [None] * len(fivehuna)           #beta
peg =  [None] * len(fivehuna)       #payout ratio
pTs =  [None] * len(fivehuna)           #PtS trailing 12 months
eps =  [None] * len(fivehuna)           #EPS
overall =  [None] * len(fivehuna)       #Overall comparision
predicted = [None] * len(fivehuna)      #Declaring necasarry arrays
current = [None] * len(fivehuna)
DiffTot = [None] * len(fivehuna)


count = len(fivehuna)

speed = 1





# Set up authentication credentials
#s = sharepy.connect("https://myuu-my.sharepoint.com/", "sam.drotar@my.uu.edu", "Notinabox3!") #auth=("sam.drotar@my.uu.edu", "Notinabox3$"))

# Get the data from a SharePoint list
#r = s.get("https://myuu-my.sharepoint.com/_api/web/lists/GetByTitle('Email S')/items")

# Parse the response as JSON
#data = r.json()

# Extract the desired column and store it as an array
#Stocked User emails from Gumroad
#userEmails = [item["user_email"] for item in data["value"]]



#@jit(target_backend='cuda') 
def get_ratings(symbols):

    comb = pd.DataFrame(columns=['Symbol', 'Predicted-Percent-Difference', 'Current-Price', 'Predicted-Price', 'Error', 'Date'])

    dow = 0
    
    while dow < count:
        try:
            predicted[dow] = stocker.predict.tomorrow(symbols[dow])  #ccomes up with the predicted price
        except KeyError:
            pass
        except TimeoutError:
            pass
        except IndexError:
            pass
        except ValueError:
            pass
        except AttributeError:
            pass
        except KeyError: 'HistoricalPriceStore'

            

        #if(dow % 20 == 0):
        print(dow)

        dow = dow + speed
    today = dt.datetime.today().date()
    dow = 0
    while dow < count:

        try:
            ticker = yf.Ticker(symbols[dow])
            todays_data = ticker.history(period='1d')
            
            #real = stocker.get_data.data.get_last_iex(symbols[dow])   #Extracts the current price 
            #current[dow] = ticker.info['regularMarketPrice']
            current[dow] = todays_data['Close'][0]

        except KeyError:
            pass
        except IndexError:
            pass

        if(dow % 20 == 0):
            print(dow)

        dow = dow + speed



    dow = 0
    while dow < count:  #Percent differance
        
        predictedPrice = 0
        error = 0
        

        try:
            servant = predicted[dow]
            servant = str(servant)
            intern = servant.split(", ")
            intern.remove(intern[2])
            intern.remove(intern[0])
            error = (intern.pop(0))
            error = float(error)
        except IndexError:
            pass

        try:
            predicted[dow] = str(predicted[dow])
            intern = predicted[dow].split(", ")
            intern.remove(intern[2])
            predictedPrice = (intern.pop(0))
            predictedPrice = predictedPrice[1:]
            predictedPrice = float(predictedPrice)
        except IndexError:
            pass

        if predictedPrice != 0:
            try:
                PercDif = ((predictedPrice - current[dow]) / current[dow] * 100)
                PercDif = round(PercDif, 3)
                DiffTot[dow] = PercDif
            except TypeError:
                pass
        else:
            DiffTot[dow] = 0


        comb = comb.append({
            'Symbol': symbols[dow],
            'Predicted-Percent-Difference': DiffTot[dow],
            'Current-Price': current[dow],
            'Predicted-Price': predictedPrice,
            'Error': error,
            'Date': today,
        }, ignore_index=True)

        predictedPrice = 0
        

        dow = dow + 1

    
         
    comb = comb.sort_values(by=['Predicted-Percent-Difference'], ascending=False)
    comb = comb.reset_index(drop=True)
        
    
    print(comb[:20])
    return comb[:20]







today = dt.datetime.today().date()

if today.isoweekday() == 6 or today.weekday() == 7:
    exit
else:
    message = get_ratings(fivehuna)

    i = 0
    while i < 20:
        try:
            error = message.at[i,'Error']
            predicted_return = message.at[i, 'Predicted-Percent-Difference']
            if error > predicted_return:
                message = message.drop([i])
                i = i + 1
            else:
                i = i + 1
        except KeyError:
            continue

    workbook = openpyxl.load_workbook('DailyOutput.xlsx')
    sheet = workbook['Sheet1']

    # get the maximum row with data
    max_row = sheet.max_row

    # find the first empty row
    for i in range(1, max_row+1):
        if sheet.cell(row=i, column=1).value is None:
            first_empty_row = i
            break
        else:
            first_empty_row = max_row + 1

    for r in dataframe_to_rows(message, index=False, header=False):
        sheet.append(r)

    workbook.save('DailyOutput.xlsx')

    message['Error'] = message['Error'].apply(lambda x: '{:.2f}%'.format(x*1))
    message['Predicted-Percent-Difference'] = message['Predicted-Percent-Difference'].apply(lambda x: '{:.2f}%'.format(x*1))
    message['Current-Price'] = message['Current-Price'].apply(lambda x: '${:.2f}'.format(x*1))
    message['Predicted-Price'] = message['Predicted-Price'].apply(lambda x: '${:.2f}'.format(x*1))

    print(message)

    smtp_server = "smtp.gmail.com"
    port = 587  # For starttls
    sender_email = "stockedaiann@gmail.com"
    free_plan =[]
    tenamonth_plan =[]
    twentyamonth_plan =["sam.drotar@gmail.com", "thatsamisam@gmail.com"]
    password = "dkowghdovlmrccfn"


    msg = MIMEMultipart()
    msg['Subject'] = "Stocked.AI"

    # Create a secure SSL context
    context = ssl.create_default_context()

    html = """
    <html>
    <head>
        <style>
          @media only screen and (min-width: 480px) {{
            table {{
                border-collapse: collapse;
                width: 60%;
                margin-left: 20%;
                margin-right: 20%;
            }}
          }}
      

          th, td {{
            border: 1px solid black;
            padding: 8px;
            text-align: left;
          }}

          th {{
            background-color: #EAE7DC;
          }}

          h2 {{
            color: black;
          }}

          @media only screen and (max-width: 480px) {{
            table {{
                border-collapse: collapse;
                max-width: 100%;
            }}
          }}
      
        </style>
    </head>
    <body>
        <h2 style="text-align: center;">Top Stock Picks for Today</h2>
        {0}
    </body>
    </html>
    """.format(message.to_html(index = False))

    part1 = MIMEText(html, 'html')
    msg.attach(part1)


    # Try to log in to server and send email
    try:
        server = smtplib.SMTP(smtp_server,port)
        server.ehlo() # Can be omitted
        server.starttls(context=context) # Secure the connection
        server.ehlo() # Can be omitted
        server.login(sender_email, password)
        server.sendmail(sender_email, twentyamonth_plan, msg.as_string())
    except Exception as e:
        # Print any error messages to stdout
        print(e)
    finally:
        server.quit() 






